from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import numbers

# ── DISCRETE PASTEL PALETTE ───────────────────────────────────────────────

FONT_COLORS = {
    0: 'FFFFFFFF', 1: 'FFFFFFFF', 2: 'FF000000',   # black text on light blues
    3: 'FF000000', 4: 'FFFFFFFF', 5: 'FFFFFFFF',   # black text on light reds
    'nan': 'FF000000'                              # black on white
}

COLORS = {
    0: 'FF053061',   # very light blue
    1: 'FF4393C3',
    2: 'FFD1E5F0',
    3: 'FFFDDBC7',
    4: 'FFD6604D',
    5: 'FF67001F',
    'nan': 'FFFFFFFF'
}

# ── PERSONA GROUPS (unchanged) ────────────────────────────────────────────
col_types = {
    'Business': ['Business Decisions'],
    'Data':     ['Data Quality', 'Open Banking (PSD2 Standard)', 'Scoring – Credit Risk',
                 'CRM Labelling', 'Solution Replacement (incl. Internal)'],
    'ESG':      ['ESG – Scope Reporting', 'ESG – CO2 Footprint'],
    'UX':       ['Transaction history', 'Mastercard Mandate', 'PFM', 'Subscription',
                 'ATM / Withdrawal', 'AI Chatbot'],
}

# ── MAIN XLSX WRITER ──────────────────────────────────────────────────────
def save_to_xlsx(df: pd.DataFrame,
                 save_path: str | Path,
                 add_separator_columns: bool = True) -> None:
    """
    Write an Excel “heat-map” version of the persona scores.

    Parameters
    ----------
    df  : DataFrame
        Must contain:
          * id | name | original_segment
          * one column per persona with scores 0–5 or NaN
          * optional "<persona> - hint" columns
    save_path : str or Path
        Target .xlsx file.
    add_separator_columns : bool, default True
        If True, inserts one narrow blank column between each persona group
        (Business ‖ Data ‖ ESG ‖ UX) for easier reading.
    """

    if df.empty:
        print("No data to write.")
        return

    # 1‣ Column order & existence check -------------------------------------
    group_lists  = [sorted(cols) for cols in col_types.values()]
    persona_cols = [c for lst in group_lists for c in lst]

    required_cols = ['id', 'name', 'original_segment'] + persona_cols
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"DataFrame is missing columns: {missing}")

    dfv = df[required_cols].copy()

    # 2‣ Workbook & header ---------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Persona scores"

    header_font  = Font(bold=True, size=9)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    col_idx = 1
    # fixed “front” columns
    for col_name in ('id', 'name', 'original_segment'):
        # Show LinkedIn URL instead of name in the second column
        if col_name == 'original_segment':
            header_text = 'original\nsegment'
        elif col_name == 'name':
            header_text = 'LinkedIn URL'
        else:
            header_text = col_name
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font, cell.alignment = header_font, header_align

        # column widths
        letter = get_column_letter(col_idx)
        if col_name == 'name':
            ws.column_dimensions[letter].width = 45
        elif col_name == 'id':
            ws.column_dimensions[letter].width = 15
        elif col_name == 'original_segment':
            ws.column_dimensions[letter].width = 10
        else:
            ws.column_dimensions[letter].width = 12

        # width = 25 if col_name == 'name' else (15 if col_name == 'id' else 14)
        # ws.column_dimensions[get_column_letter(col_idx)].width = width
        col_idx += 1

    # persona columns (+ optional separators)
    for g, cols in enumerate(group_lists):
        for persona in cols:
            cell = ws.cell(row=1, column=col_idx, value=persona)
            cell.font, cell.alignment = header_font, header_align
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
            col_idx += 1

        if add_separator_columns and g < len(group_lists) - 1:
            # insert narrow blank column
            ws.column_dimensions[get_column_letter(col_idx)].width = 3
            col_idx += 1

    ws.row_dimensions[1].height = 36
    first_persona_col_letter = get_column_letter(4)  # still column D
    ws.freeze_panes = f"{first_persona_col_letter}2"

    # 3‣ Data rows -----------------------------------------------------------
    for r, (_, row) in enumerate(dfv.iterrows(), start=2):
        # id (hyperlink if URL)
        id_val = row['id']
        id_cell = ws.cell(row=r, column=1, value=id_val)
        if isinstance(id_val, str) and id_val.startswith(('http://', 'https://')):
            id_cell.hyperlink, id_cell.style = id_val, 'Hyperlink'

        # name column now displays LinkedIn URL if available
        if isinstance(id_val, str) and id_val.startswith(('http://', 'https://')):
            display_val = id_val
        else:
            display_val = row['name']

        name_cell = ws.cell(row=r, column=2, value=display_val)
        if isinstance(display_val, str) and display_val.startswith(('http://', 'https://')):
            name_cell.hyperlink, name_cell.style = display_val, 'Hyperlink'

        # original_segment
        ws.cell(row=r, column=3, value=row['original_segment'])

        c = 4
        for gi, cols in enumerate(group_lists):
            for persona in cols:
                score = row[persona]
                if pd.isna(score) or not isinstance(score, numbers.Number):
                    key = 'nan'
                else:
                    key = int(score)
                bg_hex, font_hex = COLORS[key], FONT_COLORS[key]

                cell = ws.cell(row=r, column=c, value=None if key == 'nan' else key)
                cell.fill  = PatternFill('solid', start_color=bg_hex, end_color=bg_hex)
                cell.font  = Font(color=font_hex)
                cell.alignment = Alignment(horizontal='center', vertical='center')

                hint_col = f"{persona} - hint"
                if hint_col in df and pd.notna(df.at[row.name, hint_col]):
                    com = Comment(str(df.at[row.name, hint_col]), "GPT-Note",)
                    com.auto_size = True  # let Excel decide how big
                    cell.comment = com
                c += 1
            if add_separator_columns and gi < len(group_lists) - 1:
                c += 1  # skip separator

    # 4‣ Save ---------------------------------------------------------------
    Path(save_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    print(f"Wrote {save_path}")
